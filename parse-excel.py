import openpyxl
import json

wb = openpyxl.load_workbook('C:/Users/saarm/Downloads/Clark Rd - EmcoList Price (1).xlsx')
ws = wb.active

packages = {
    'Forecourt Island Equipment': [],
    'Forecourt Submerged Pump Package': [],
    'Tank Equipment': []
}

current_package = None

for i in range(2, ws.max_row + 1):
    row = ws[i]
    part_num = row[0].value
    desc = row[1].value
    list_price = row[2].value
    qty = row[3].value
    
    # Check if Column A has a package name
    if part_num and str(part_num) in packages:
        current_package = str(part_num)
        continue
    
    # Skip if no part number or no description
    if not part_num or not desc or qty == 'Total:':
        continue
    
    # Add item to current package
    if current_package and list_price is not None and qty is not None:
        try:
            packages[current_package].append({
                'partNum': str(part_num),
                'desc': str(desc),
                'cost': float(list_price),
                'qty': float(qty)
            })
        except (ValueError, TypeError):
            pass

print(json.dumps(packages, indent=2))
